'''
Problem Statement:
1. Please import csv (or) openpyxl & Logging packages for this problem.
2. Program should take any filename as per the format mentioned as input.
3. Input month value from the file name where  eg :january(expedia_report_monthly_january_2018.xlsx)
4. Based on the month and year input values value print the values in logfile using logger

This is from the first tab :

Eg for January :

Calls Offered: 16,915
Abandon after 30s : 2.32%
FCR : 86.50%
DSAT :  14.20%
CSAT : 78.30%

Similarly go to "VOC Rolling MoM" tab

Grab all the values related to Jan-18 and print.

In Net Promoter Score : Promoters => 200 : good Promoters <200 : bad
			Passives => 100 : good Passives <100 : bad
			Decractors => 100 : good Decrators <100 : bad

Rest all values remain the same.

'''


from datetime import datetime
from openpyxl import load_workbook
import os
import logging
import webbrowser


def list_excel_files(file_path):
    """Will list out all Excel (.xls) files in the given file path and will check if there are no .xls files in the path and exit the script

    Args:
        file_path ([string]): absolute path to the location of Excel files to analzye

    Returns:
        excel_files ([list]): a list of the available Excel files
    """
    # Changing directory to the folder where my Excel files are
    os.chdir(file_path)

    # Listing what's in the folder I selected
    lst = os.listdir()
    logging.info(f"Searching all .xls files in {os.path.basename(file_path)}")

    # Retrieve only .xls files (Excel files)
    excel_files = [s for s in lst if '.xls' in s]
    if not excel_files:
        logging.warning(
            f"There are no Excel Files in {os.path.basename(file_path)}!")
        print(
            f"There are no Excel files in {os.path.basename(file_path)}. Please check the file path and rerun this script")
        exit()
    return excel_files


def choose_excel_file(excel_files):
    """Given a list of Excel files, this function will iterate thru the list and return the user chosen Excel File and assign it to excel_file.
    The function will also check if the user input is valid (User input should be an integer and in the list itself)


    Args:
        excel_files ([list]): a list of the available Excel files

    Returns:
        excel_file [string]: a string of the full Excel file name (eg. expedia_report_monthly_january_2018.xls)
    """
    logging.info("Asking user for what file to analyze")
    while(True):
        try:
            print("These Excel Files came up:")
            for index, file in enumerate(excel_files, start=1):
                print(f"{index}. {file}")
            choice = input("\nSelect the number next to the file you'd like: ")
            excel_file = excel_files[int(choice) - 1]
            if choice.isdigit() and int(choice) <= len(excel_files):
                logging.info(f"User chose {excel_file}")
                return excel_file

        except Exception:
            print("ERROR: You must choose a valid option\n")
            logging.warning("User selected a non-valid option")


def first_worksheet_data(row_data, header_1st_sheet, wb):
    """Will iterate thru row_data and header_1st_sheet to output to the log of this script (wb)
    This particular set of data had one data that was NOT a float and the rest were so, the for loop accounts for the percentages/floats

    Args:
        row_data ([list]): list of integers and floats from our 1st worksheet
        header_1st_sheet ([list]): list of strings in the 1st cells of each row in the 1st worksheet
        wb ([class]): the Excel file workbook
    """
    # Outputting data from the 1st worksheet

    logging.info(f"Data from {wb.worksheets[0]}")
    logging.info(f"{header_1st_sheet[0]}: {row_data[1]}")
    for index in range(1, len(header_1st_sheet)):
        logging.info(
            f"{header_1st_sheet[index].strip()}: {row_data[index + 1]:.2%}")


def second_worksheet_data(column_data, header_2nd_sheet, wb):
    """Iterating thru column data and header_2nd_sheet to output to the log of this script (wb)
    This function will also check for the conditionals given to us by the problem statement shown down below:
            Conditionals:
            Promoters => 200 : good     Promoters <200 : bad
            Passives => 100 : good      Passives <100 : bad
            Decractors => 100 : good    Decrators <100 : bad
    Args:
        column_data ([list]): list of integers and floats from the 2nd worksheet
        header_2nd_sheet ([list]): list of strings in the 1st cells of each column in the 2nd worksheet
        wb ([class]): the Excel file workbook
    """
    # Outputting data from our 2nd worksheet

    logging.info(f"Data from {wb.worksheets[1]}")
    for index in range(0, len(column_data)):
        if type(column_data[index]) == float:
            logging.info(
                f"{header_2nd_sheet[index+1].strip()}: {column_data[index]:.2%}")

        elif header_2nd_sheet[index+1] == "Promoters (Recommend Score 9 to 10)":
            if column_data[index] >= 200:
                logging.info(
                    f"{header_2nd_sheet[index+1].strip()}: {column_data[index]}, which is good!")
            else:
                logging.warning(
                    f"{header_2nd_sheet[index+1].strip()}: {column_data[index]}, which is bad!")

        elif header_2nd_sheet[index + 1] == "Passives (Recommend Score 7 to 8)" or header_2nd_sheet[index + 1] == "Dectractors (recommend Score 0 to 6)":
            if column_data[index] >= 100:
                logging.info(
                    f"{header_2nd_sheet[index+1].strip()}: {column_data[index]}, which is good!")
            else:
                logging.warning(
                    f"{header_2nd_sheet[index+1].strip()}: {column_data[index]}, which is bad!")

        else:
            logging.info(
                f"{header_2nd_sheet[index+1].strip()}: {column_data[index]}")


def open_log_chrome(file_name):
    """Will open the webbrowser, Chrome, to the given file_name.

    Args:
        file_name ([string): full file name with extension as well (eg. test.log)
    """
    chrome_path = "C:/Program Files/Google/Chrome/Application/chrome.exe %s"
    answer = input(
        "The log file has been created/edited. Would you like to open it?\n")

    try:
        if answer.lower() == "yes" or answer.lower() == "y":
            webbrowser.get(chrome_path).open(
                'file://' + os.path.realpath(f'{file_name}'))

        if answer.lower() == "no" or answer.lower() == "n":
            print("Thank you for your time")
    except Exception:
        print("Please type in a valid answer")
    finally:
        print("Have a great day!")


if __name__ == "__main__":
    # Creating an instance of logging library
    logging_file_name = "test.log"
    logging.basicConfig(filename=f"C:\\Users\\Eric\Documents\\Python Lessons\\SmoothStack\\week_1_excel\\{logging_file_name}", level=logging.INFO,
                        format="%(asctime)s %(levelname)s: %(message)s")

    # File path to the excel files you wish to analyze
    file_path = 'C:\\Users\\Eric\\Documents\\Python Lessons\\SmoothStack\\week_1_excel'
    # Edit file path - NO HARDCODING, should run regardless of where the script is

    excel_files = list_excel_files(file_path)
    chosen_file = choose_excel_file(excel_files)
    chosen_file_path = os.path.abspath(chosen_file)

    # Loading Excel file for file chosen by user
    logging.info("Loading Excel File user chose")
    wb = load_workbook(chosen_file_path)

    # Create variables for worksheets 1 and 2
    ws_summary = wb.worksheets[0]
    ws_voc = wb.worksheets[1]

    # Retrieve Month and Year to filter our data from our excel spreadsheet
    logging.info(
        "Grabbing month and year from the file name to filter the data")

    # Get the 1st element of excel_files and seperate each word + get rid of .xslx extension
    month_year = (os.path.splitext(chosen_file)[0].split("_"))[-2:]

    # Combining two string (month + year) together and using datetime module to match the Excel data
    month_year_split = ' '.join(month_year).title()
    month_year_formatted = datetime.strptime(month_year_split, "%B %Y")

    # Capitalizing month to work with the conditional statements later down our codebase
    month_year[0] = month_year[0].capitalize()

    # Extracting our columns and rows to iterate through and get the selected data we want
    worksheet_rows_1st_sheet = list(ws_summary.iter_rows(values_only=True))
    worksheet_columns_1st_sheet = list(ws_summary.iter_cols(values_only=True))
    worksheet_rows_2nd_sheet = list(ws_voc.iter_rows(values_only=True))

    # Iterating through each row in 1st worksheet and checking for the Month + Year we want
    # Will only grab the row of data if it's not EMPTY
    for row in worksheet_rows_1st_sheet:
        if row[0] == month_year_formatted:
            row_data = [row for row in row if row != None]

    # Getting header data and not grabbing the EMPTY cells
    logging.info("Getting header data from 1st sheet")
    header_1st_sheet = [col[0]
                        for col in worksheet_columns_1st_sheet if col[0] != None]
    header_2nd_sheet = [col[0]
                        for col in worksheet_rows_2nd_sheet if col[0] != None and col[0] != "Overall NPS %"]

    # Generating data (col_data) from 2nd worksheet's column only if it's equal to the Month + Year or Month only
    worksheet_columns = list(ws_voc.iter_cols(values_only=True))
    for col in worksheet_columns:
        if col[0] == month_year_formatted:
            col_data = [col for col in col if col != None]

        else:
            # Will search by getting the month's full name eg. January or March
            if col[0] == month_year[0]:
                logging.warning(
                    "This column does NOT have a year! Please add it for consistency")
                col_data = [col for col in col if col != None]

    # Some categories have 2 values, a percentage and whole number. In this scenario, we only want the whole number.
    # Removing the percentage values from our list of data
    for column in col_data[0:8]:
        if type(column) != int:
            col_data.remove(column)

    first_worksheet_data(row_data, header_1st_sheet, wb)
    second_worksheet_data(col_data, header_2nd_sheet, wb)
    open_log_chrome(logging_file_name)
