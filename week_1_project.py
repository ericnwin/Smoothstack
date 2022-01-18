# Author: Eric Nguyen

'''
Problem Statement:
1. Please import csv (or) openpyxl & Logging packages for this problem.
2. Program should take any filename as per the format mentioned as input.
3. Input month value from the file name where  eg :january(expedia_report_monthly_january_2018.xlsx)
4. Based on the month and year input values value print the values in logfile using logger

This is from the first tab :

Eg for January :

Calls Offered: 16,915
Abandon after 30s : 2.32%
FCR : 86.50%
DSAT :  14.20%
CSAT : 78.30%

Similarly go to "VOC Rolling MoM" tab

Grab all the values related to Jan-18 and print.

In Net Promoter Score : Promoters => 200 : good Promoters <200 : bad
			Passives => 100 : good Passives <100 : bad
			Decractors => 100 : good Decrators <100 : bad

Rest all values remain the same.

'''
from datetime import datetime
from openpyxl import load_workbook
import os
import logging

# Creating an instance of logging library
logging.basicConfig(filename=r"C:\Users\Eric\Documents\Python Lessons\SmoothStack\week_1_excel\test.log", level=logging.INFO,
                    format="%(asctime)s %(levelname)s: %(message)s")

# Changing directory to the folder where my Excel files are
os.chdir(r'C:\Users\Eric\Documents\Python Lessons\SmoothStack\week_1_excel')

# Listing what's in the folder I selected
lst = os.listdir()
logging.info("Searching all .xls files in week_1_excel folder")

# Retrieve only .xls files (Excel files)
logging.info("Asking user for what file to analyze")
excel_files = [s for s in lst if '.xls' in s]
while(True):
    for index, file in enumerate(excel_files, start=1):
        print(f"{index}. {file}")
    choice = input("\nSelect the number next to the file you'd like: ")
    if choice.isdigit() and int(choice) <= len(excel_files):
        break
    else:
        print("ERROR: You must choose a valid option\n")
        logging.warning("User selected a non-valid option")

excel_file = excel_files[int(choice) - 1]
logging.info(f"User chose {excel_file}")
file_path = os.path.abspath(excel_file)

# Loading Excel file for file chosen by user
logging.info("Loading Excel File user chose")
wb = load_workbook(file_path)

# Create variables for worksheets 1 and 2
ws_summary = wb.worksheets[0]
ws_voc = wb.worksheets[1]


# Get the 1st element of excel_files and seperate each word + get rid of .xslx extension
print(os.path.splitext(excel_file)[0].split("_"))

# Retrieve Month and Year to filter our data from our excel spreadsheet
logging.info("Grabbing month and year from the file name to filter the data")
month_year = (os.path.splitext(excel_file)[0].split("_"))[-2:]

month_year_split = ' '.join(month_year).title()
month_year[0] = month_year[0].capitalize()
month_year_formatted = datetime.strptime(month_year_split, "%B %Y")
print(month_year_formatted)

# Convert string "January 2018" to actual date format
# The cell value is actually a date (1/1/2018) along with a time of 00:00:00, hence the necessity of this step

worksheet_rows_1st_sheet = list(ws_summary.iter_rows(values_only=True))
worksheet_columns_1st_sheet = list(ws_summary.iter_cols(values_only=True))
worksheet_rows_2nd_sheet = list(ws_voc.iter_rows(values_only=True))

for row in worksheet_rows_1st_sheet:
    if row[0] == month_year_formatted:
        row_data = [row for row in row if row != None]
        print(row_data)

# Getting header data
logging.info("Getting header data from 1st sheet")
header_1st_sheet = [col[0]
                    for col in worksheet_columns_1st_sheet if col[0] != None]
header_2nd_sheet = [col[0]
                    for col in worksheet_rows_2nd_sheet if col[0] != None and col[0] != "Overall NPS %"]
worksheet_columns = list(ws_voc.iter_cols(values_only=True))
print(header_2nd_sheet)

print(f"Printing month_year: {month_year[0]}")
for col in worksheet_columns:
    if col[0] == month_year_formatted:
        col_data = [col for col in col if col != None]
        print(col_data)

    else:
        # Will search by getting the month's full name eg. January or March
        if col[0] == month_year[0]:
            col_data = [col for col in col if col != None]


logging.info(f"Data from {wb.worksheets[0]}")
logging.info(f"{header_1st_sheet[0]}: {row_data[1]}")
for index in range(1, len(header_1st_sheet)):
    logging.info(f"{header_1st_sheet[index].strip()}: {row_data[index+1]:.2%}")


# Some categories have 2 values, a percentage and whole number. In this scenario, we only want the whole number.
# Removing the percentage values from our list of data
for column in col_data[0:8]:
    if type(column) != int:
        col_data.remove(column)

logging.info(f"Data from {wb.worksheets[1]}")
for index in range(0, len(col_data)):
    if type(col_data[index]) == float:
        logging.info(
            f"{header_2nd_sheet[index+1].strip()}: {col_data[index]:.2%}")
        ''' Conditionals:
        Promoters => 200 : good     Promoters <200 : bad
		Passives => 100 : good      Passives <100 : bad
		Decractors => 100 : good    Decrators <100 : bad
        '''
    elif header_2nd_sheet[index+1] == "Promoters (Recommend Score 9 to 10)":
        if col_data[index] >= 200:
            logging.info(
                f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}, which is good!")
        else:
            logging.warning(
                f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}, which is bad!")
    elif header_2nd_sheet[index + 1] == "Passives (Recommend Score 7 to 8)" or header_2nd_sheet[index + 1] == "Dectractors (recommend Score 0 to 6)":
        if col_data[index] >= 100:
            logging.info(
                f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}, which is good!")
        else:
            logging.warning(
                f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}, which is bad!")
    else:
        logging.info(f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}")


def second_worksheet_data(col_data, header_2nd_sheet):
    for index in range(0, len(col_data)):
        if type(col_data[index]) == float:
            logging.info(
                f"{header_2nd_sheet[index+1].strip()}: {col_data[index]:.2%}")
            ''' Conditionals:
            Promoters => 200 : good     Promoters <200 : bad
            Passives => 100 : good      Passives <100 : bad
            Decractors => 100 : good    Decrators <100 : bad
            '''
        elif header_2nd_sheet[index+1] == "Promoters (Recommend Score 9 to 10)":
            if col_data[index] >= 200:
                logging.info(
                    f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}, which is good!")
            else:
                logging.warning(
                    f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}, which is bad!")
        elif header_2nd_sheet[index + 1] == "Passives (Recommend Score 7 to 8)" or header_2nd_sheet[index + 1] == "Dectractors (recommend Score 0 to 6)":
            if col_data[index] >= 100:
                logging.info(
                    f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}, which is good!")
            else:
                logging.warning(
                    f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}, which is bad!")
        else:
            logging.info(
                f"{header_2nd_sheet[index+1].strip()}: {col_data[index]}")
